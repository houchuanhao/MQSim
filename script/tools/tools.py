import xml.etree.ElementTree as ET
import openpyxl
import re
import os
xml_ssdcfg='ssdconfig.xml'
xml_workload='workload.xml'
xml_out='workload_scenario_1.xml'
path_sensitive= 'sensitive.xlsx'
path_readme = '../../README.md'
precision = ['Overprovisioning_Ratio','GC_Exect_Threshold','GC_Hard_Threshold','PCIe_Lane_Bandwidth']
xlsx_config = "config.xlsx"

def xlsx2lst(path,sheetname=None):
    lst =[]
    workbook = openpyxl.load_workbook(path)
    sheet = None
    if sheetname==None:
        sheet = workbook.active
    else:
        sheet = workbook[sheetname]
    for row in sheet.iter_rows(values_only=True):
        lst_row = list(row)
        for i in range(len(lst_row),6+1):
            lst_row.append(None)
        lst.append(lst_row)
    #print(lst)
    return lst
def is_numeric(string):
    int_pattern = r'^[-+]?\d+$'  # 整数模式 返回1
    float_pattern = r'^[-+]?[0-9]*\.?[0-9]+([eE][-+]?[0-9]+)?$'  # 浮点数模式 返回2
    if bool(re.match(int_pattern, string)):
        return 1
    if bool(re.match(float_pattern, string)):
        return 2
    return 0


def travel(root,l=0):
    s = ""
    for i in range(l):
        s = s + "   "
    for child in root:
        if len(child)>0:
            print(s, child.tag)
            travel(child,l+1)
        else:
            print(s,child.tag,":  ",child.text)
def getext(root,key):
    for child in root:
        if len(child)>0:
            t = getext(child,key)
            if t!=None:
                return t
        else:
            if(child.tag==key):
                return child.text
    return None
            #print(s,child.tag,":  ",child.text)
    # 读取XML文件
def setext(root,key,value):
    for child in root:
        if len(child)>0:
            t = setext(child,key,value)
            if t!=None:
                return 1
        else:
            if(child.tag==key):
                child.text = value
                return 1
    return None

#travel(root)
def root2dic(root, dic=None):
    if dic==None:
        dic = {}
    #print(dic)
    for child in root:
        if child.tag in dic:
            print("error")
        if len(child)>0:
            dic[child.tag] = None
            dic = root2dic(child, dic) # 可变传参 可以优化
        else:
            dic[child.tag] = child.text
    return dic
def dic2root(dic,root):
    for key,value in dic.items():
        setext(root,key,str(value))
def dic2xlsx(dic,path=path_sensitive):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    i = 1
    for key,value in dic.items():
        v = value
        if value != None:
            if(is_numeric(value)==1):
                v = int(value)
            if (is_numeric(value) == 2):
                v = float(value)
        worksheet['A' + str(i)] = key
        worksheet['B' + str(i)] = v

        i = i + 1
    workbook.save(path)
    workbook.close()

def saveTree(tree,path):
    tree.write(path)
def getTree(path):
    tree = ET.parse(path)
    root = tree.getroot()
    return tree,root
def lst2excel(Lst,path):
    print("lst2excel "+path)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None
    for i, lst in enumerate(Lst, start=1):
        for j,value in enumerate(lst,start=1):
            #print("i="+str(i)," j="+str(j),"value= ",value)
            worksheet.cell(row=i, column=j).value = str(value)
    #print(path)
    workbook.save(path)
def lst2sheet(path,sheet_name,Lst):
    workbook = openpyxl.Workbook()
    sheet_names = workbook.sheetnames # 先删除
    if sheet_name in sheet_names:
        workbook.remove(workbook[sheet_name])
        print("remove "+sheet_name)
    worksheet = workbook.create_sheet(title=sheet_name+"zzz")  # 新建sheet
    worksheet = workbook.create_sheet(title=sheet_name) # 新建sheet
    for i, lst in enumerate(Lst, start=1):
        for j,value in enumerate(lst,start=1):
            worksheet.cell(row=i, column=j).value = value
    print(workbook.sheetnames)
    workbook.save(path)
    workbook.close()
def xml2dic(ssd,workload):
    #print(ssd," ",workload)
    tree , root = getTree(ssd)
    dic_ssd = root2dic(root)
    tree , root = getTree(workload)
    dic_workload = root2dic(root)

    return dic_ssd,dic_workload

def str2num(s):
    if is_numeric(s)==0:
        return s
    if is_numeric(s)==1:
        return int(s)
    if is_numeric(s)==2:
        return float(s)

def lst2dic(lst, p=6):
    dic = {}
    for data in lst:
        dic[data[0]] = data[p]

    return dic

def get_all_folders(path):
    folders = []
    for item in os.listdir(path):
        item_path = os.path.join(path, item)
        if os.path.isdir(item_path):
            folders.append(item_path)
    return folders


def main():
    keys = []
    range = []
    ssdsheet = "ssd0"
    workloadsheet = "workload0"
    workspace = "workspace_" + ssdsheet + "_" + workloadsheet
    # tree_ssd, root_ssd = tools.getTree(tools.path_ssd)
    lst_ssd = xlsx2lst(xlsx_config, ssdsheet)
    lst_workload = xlsx2lst(xlsx_config, workloadsheet)
    print(lst_ssd,lst_workload)
    print("???")
#main()